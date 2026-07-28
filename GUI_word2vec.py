"""
한국어 문맥 연관어 분석기 - Word2Vec/PMI 다중 TXT 파일 GUI 버전
============================================================

이 프로그램의 목적
------------------
1. 사용자가 여러 개의 뉴스 TXT 파일을 한 번에 선택한다.
2. 선택한 모든 파일의 문장을 하나의 말뭉치(corpus)로 합친다.
3. Kiwi 형태소 분석기로 의미가 있는 단어를 추출한다.
4. Word2Vec 또는 PMI로 단어 사이의 문맥 연관성을 계산한다.
5. MZ, 4050, 기성, 청년처럼 서로 비슷한 위치에서 사용되는 단어를 찾는다.
6. 분석 결과와 원본 문장을 Excel 파일로 저장한다.

중요한 분석 원리
----------------
- 기준어가 들어 있는 문장만 골라서 분석하지 않는다.
- 선택한 파일의 모든 문장을 학습한다.
- 기준어는 학습이 끝난 뒤 유사 단어를 조회할 때 사용한다.
- MZ / MZ세대 / MZ 세대처럼 같은 의미의 표기는 하나의 대표어로 통합한다.
- '세대', '실태', '관련'처럼 기준어와 자주 붙지만 비교 대상은 아닌 단어는
  결과 제외어로 설정할 수 있다.

필요한 외부 라이브러리
----------------------
pip install gensim kiwipiepy numpy openpyxl matplotlib

Python 기본 라이브러리인 tkinter는 보통 Python 설치 시 함께 설치된다.
"""

from __future__ import annotations

# os: CPU 코어 수 확인, Windows에서 결과 폴더 열기 등에 사용한다.
import os

# math: PMI 점수 계산의 로그 연산에 사용한다.
import math

# queue: Word2Vec 학습 스레드에서 GUI 스레드로 로그와 완료 신호를 안전하게 전달한다.
import queue

# re: 표현 통합, 파일명 정리, 토큰 분리 등에 사용하는 정규표현식 모듈이다.
import re

# subprocess: macOS/Linux에서 결과 폴더를 여는 명령을 실행한다.
import subprocess

# sys: 현재 운영체제 확인, 실행 환경 관련 처리에 사용한다.
import sys

# threading: 분석 중 GUI 창이 멈추지 않도록 학습을 별도 스레드에서 실행한다.
import threading

# Counter: 각 단어가 어떤 품사로 분석되었는지 횟수를 기록한다.
# defaultdict: 단어별 Counter를 자동 생성하기 위해 사용한다.
from collections import Counter, defaultdict

# datetime: 결과 파일명과 Excel 메타데이터에 생성 시간을 기록한다.
from datetime import datetime

# Path: Windows/Unix 경로를 안정적으로 다루기 위해 사용한다.
from pathlib import Path

# Callable: 로그 함수 등의 타입 힌트를 작성하기 위해 사용한다.
from typing import Callable

# tkinter: Python 기본 GUI 라이브러리다.
import tkinter as tk

# filedialog: 파일 선택/저장 창
# messagebox: 오류·완료 메시지 창
# scrolledtext: 스크롤 가능한 로그 출력 창
# ttk: 기본 tkinter보다 현대적인 모양의 위젯
from tkinter import filedialog, messagebox, scrolledtext, ttk


# =============================================================================
# 1. 분석 기본 설정
# =============================================================================

# Word2Vec 학습에 포함할 Kiwi 품사 태그다.
#
# NNG: 일반 명사      예) 소비, 문화, 청년
# NNP: 고유 명사      예) 서울, 삼성, MZ
# SL : 외국어         예) MZ, AI, ESG
# SN : 숫자           예) 2030, 4050
# VV : 동사           예) 증가하, 변화하
# VA : 형용사         예) 크, 새롭
# MAG: 일반 부사      예) 매우, 크게
#
# 조사(JKS 등), 어미(EF 등)는 단어 의미보다는 문법 기능이 강하므로 제외한다.
TRAIN_POS = {"NNG", "NNP", "SL", "SN", "VV", "VA", "MAG"}

# GUI에서 '명사 중심'을 선택했을 때 결과 후보로 허용할 품사다.
# Word2Vec 학습은 위 TRAIN_POS 전체를 사용하지만,
# 결과 출력 단계에서만 명사·영문·숫자로 제한한다.
NOUN_RESULT_POS = {"NNG", "NNP", "SL", "SN"}

# 전체 토큰 수가 이 값 이하이면 PMI, 더 크면 Word2Vec을 추천한다.
# 이 기준값과 실제 토큰 수는 사용자 화면이나 결과 파일에 표시하지 않는다.
PMI_RECOMMENDED_MAX_TOKENS = 100_000


def recommend_analysis_mode(total_token_count: int) -> str:
    """전체 토큰 수를 화면에 노출하지 않고 적합한 분석 모드만 반환한다."""

    if total_token_count <= PMI_RECOMMENDED_MAX_TOKENS:
        return "PMI"

    return "Word2Vec"


# 처음 실행했을 때 GUI에 미리 표시되는 표현 통합 규칙이다.
# 형식: 대표어=표현1|표현2; 다른대표어=표현1|표현2
#
# 예를 들어 원문에 MZ세대, MZ 세대, 엠지세대가 섞여 있어도
# 모두 'MZ'라는 하나의 토큰으로 바꿔 학습한다.
DEFAULT_ALIAS_RULES = (
    "MZ=MZ세대|MZ 세대|엠지세대|엠지 세대|엠지; "
    "2030=2030세대|2030 세대; "
    "4050=4050세대|4050 세대; "
    "기성=기성세대|기성 세대; "
    "청년=청년세대|청년 세대; "
    "중장년=중장년층|중장년세대|중장년 세대"
)

# 처음 실행했을 때 유사어 결과에서 제외할 기본 단어들이다.
# 이 단어들은 기준어와 자주 같이 등장할 수 있지만,
# 비교 집단이나 의미상 대체어로 보기 어려운 경우가 많다.
DEFAULT_EXCLUDED_WORDS = "세대, 실태, 관련, 대상, 분석, 조사, 현황"


# =============================================================================
# 2. 사용자 단어와 표현 통합 규칙 처리
# =============================================================================

def parse_user_word(spec: str) -> tuple[str, str, float]:
    """
    사용자 단어 문자열을 Kiwi 등록 형식으로 분리한다.

    지원 형식
    ---------
    MZ
        -> 단어=MZ, 품사=NNP, 점수=0.0

    MZ:NNP
        -> 단어=MZ, 품사=NNP, 점수=0.0

    MZ:NNP:5.0
        -> 단어=MZ, 품사=NNP, 점수=5.0

    점수가 높을수록 Kiwi가 해당 표현을 하나의 단어로 우선 인식할 가능성이 높아진다.
    """

    # ':'를 기준으로 문자열을 최대 세 부분으로 나눈다.
    parts = [part.strip() for part in spec.split(":")]

    # 첫 번째 부분은 반드시 단어여야 한다.
    word = parts[0]
    if not word:
        raise ValueError("사용자 단어에 빈 항목이 있습니다.")

    # 품사가 비어 있으면 고유명사(NNP)를 기본값으로 사용한다.
    tag = parts[1] if len(parts) > 1 and parts[1] else "NNP"

    # 점수가 비어 있으면 0.0을 사용한다.
    try:
        score = float(parts[2]) if len(parts) > 2 and parts[2] else 0.0
    except ValueError as exc:
        raise ValueError(f"사용자 단어 점수가 숫자가 아닙니다: {spec}") from exc

    return word, tag, score


def parse_alias_rules(text: str) -> list[tuple[str, list[str]]]:
    """
    GUI에서 입력한 표현 통합 규칙을 파싱한다.

    입력 예시
    ---------
    MZ=MZ세대|MZ 세대|엠지세대; 4050=4050세대|4050 세대

    반환 예시
    ---------
    [
        ("MZ", ["MZ세대", "MZ 세대", "엠지세대"]),
        ("4050", ["4050세대", "4050 세대"]),
    ]
    """

    rules: list[tuple[str, list[str]]] = []

    # 입력란을 비워 두면 아무 규칙도 적용하지 않는다.
    if not text.strip():
        return rules

    # 세미콜론 또는 줄바꿈을 규칙 구분자로 사용한다.
    chunks = [
        chunk.strip()
        for chunk in re.split(r"[;\n]+", text)
        if chunk.strip()
    ]

    for chunk in chunks:
        # 각 규칙은 대표어=표현목록 형식이어야 한다.
        if "=" not in chunk:
            raise ValueError(
                "표현 통합 규칙은 '대표어=표현1|표현2' 형식이어야 합니다: "
                + chunk
            )

        # 첫 번째 '='만 기준으로 대표어와 표현 목록을 분리한다.
        canonical, aliases_text = chunk.split("=", 1)
        canonical = canonical.strip()

        if not canonical:
            raise ValueError("표현 통합 규칙의 대표어가 비어 있습니다.")

        # 같은 대표어로 통합할 표현들은 '|'로 구분한다.
        aliases = [
            alias.strip()
            for alias in aliases_text.split("|")
            if alias.strip()
        ]

        if not aliases:
            raise ValueError(f"'{canonical}' 규칙에 통합할 표현이 없습니다.")

        rules.append((canonical, aliases))

    return rules


def normalize_text(
    text: str,
    alias_rules: list[tuple[str, list[str]]],
) -> str:
    """
    문장 안의 다양한 표기를 대표어로 통합한다.

    예시
    ----
    규칙: MZ=MZ세대|MZ 세대|엠지세대

    입력: "MZ 세대와 엠지세대의 소비 성향"
    출력: "MZ와 MZ의 소비 성향"

    긴 표현부터 먼저 바꾸는 이유
    -----------------------------
    'MZ세대'를 바꾸기 전에 'MZ'를 먼저 바꾸면 일부 문자열이 잘못 치환될 수 있다.
    따라서 긴 표현을 먼저 처리한다.
    """

    replacements: list[tuple[str, str]] = []

    for canonical, aliases in alias_rules:
        for alias in aliases:
            # 대표어 자체와 완전히 같은 표현은 굳이 치환 목록에 넣지 않는다.
            if alias.casefold() != canonical.casefold():
                replacements.append((alias, canonical))

    # 문자열 길이가 긴 표현부터 치환한다.
    replacements.sort(key=lambda item: len(item[0]), reverse=True)

    normalized = text

    for alias, canonical in replacements:
        # "MZ 세대" 규칙은 원문에서 "MZ세대", "MZ  세대"처럼
        # 공백 개수가 달라도 찾을 수 있도록 \s* 패턴을 사용한다.
        parts = [re.escape(part) for part in alias.split()]
        pattern = r"\s*".join(parts) if len(parts) > 1 else re.escape(alias)

        # 영문 대소문자를 구분하지 않고 치환한다.
        normalized = re.sub(
            pattern,
            canonical,
            normalized,
            flags=re.IGNORECASE,
        )

    return normalized


def normalize_term(
    term: str,
    alias_rules: list[tuple[str, list[str]]],
) -> str:
    """
    기준어 또는 제외어 한 개를 대표어로 통합한다.

    문장 전체가 아닌 '단어 하나'를 처리할 때 사용한다.

    예시
    ----
    사용자가 기준어에 MZ세대를 입력했더라도,
    통합 규칙에 따라 실제 조회 기준어는 MZ가 된다.
    """

    stripped = term.strip()

    # 비교할 때는 공백과 영문 대소문자를 무시한다.
    folded = re.sub(r"\s+", "", stripped).casefold()

    for canonical, aliases in alias_rules:
        candidates = [canonical, *aliases]

        for candidate in candidates:
            candidate_folded = re.sub(r"\s+", "", candidate).casefold()

            if candidate_folded == folded:
                return canonical

    # 어떤 통합 규칙에도 해당하지 않으면 원래 단어를 반환한다.
    return stripped


def dominant_pos(
    word: str,
    token_pos_counts: dict[str, Counter],
) -> str | None:
    """
    특정 단어가 가장 자주 분석된 품사를 반환한다.

    동일한 표기의 단어가 문장에 따라 서로 다른 품사로 분석될 수 있다.
    이 함수는 그중 가장 많이 등장한 품사를 대표 품사로 사용한다.
    """

    counts = token_pos_counts.get(word)

    if not counts:
        return None

    # Counter.most_common(1)은 가장 빈도가 높은 (품사, 횟수) 한 개를 반환한다.
    return counts.most_common(1)[0][0]


# =============================================================================
# 3. 입력 파일 읽기
# =============================================================================

def read_text_file(path: Path) -> tuple[str, str]:
    """
    TXT 파일을 UTF-8 우선, 실패하면 CP949로 읽는다.

    반환값
    ------
    (파일 내용, 사용한 인코딩 이름)
    """

    try:
        return path.read_text(encoding="utf-8-sig"), "UTF-8"
    except UnicodeDecodeError:
        try:
            return path.read_text(encoding="cp949"), "CP949"
        except UnicodeDecodeError as exc:
            raise ValueError(
                f"'{path.name}' 파일을 UTF-8 또는 CP949로 읽을 수 없습니다."
            ) from exc


def load_sentences(
    paths: list[str],
    log: Callable[[str], None],
) -> tuple[list[str], list[dict[str, object]]]:
    """
    여러 TXT 파일을 읽어서 하나의 학습 문장 목록으로 합친다.

    반환값
    ------
    sentences
        Word2Vec 학습에 사용할 전체 문장 리스트

    sentence_records
        Excel 원문 시트에 저장할 부가정보 리스트
        각 항목은 파일명, 파일 내 줄 번호, 문장을 포함한다.
    """

    if not paths:
        raise ValueError("학습할 TXT 파일을 한 개 이상 선택하세요.")

    all_sentences: list[str] = []
    sentence_records: list[dict[str, object]] = []

    for path_string in paths:
        path = Path(path_string).expanduser()

        if not path.exists():
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {path}")

        if not path.is_file():
            raise ValueError(f"선택한 경로가 파일이 아닙니다: {path}")

        text, encoding = read_text_file(path)

        # 비어 있는 줄은 학습에 사용하지 않는다.
        file_sentence_count = 0

        for line_number, line in enumerate(text.splitlines(), start=1):
            sentence = line.strip()

            if not sentence:
                continue

            all_sentences.append(sentence)
            file_sentence_count += 1

            # 나중에 Excel에서 문장의 출처 파일을 확인할 수 있도록 기록한다.
            sentence_records.append(
                {
                    "파일명": path.name,
                    "줄번호": line_number,
                    "문장": sentence,
                }
            )

        log(
            f"파일 읽기 완료: {path.name} "
            f"({encoding}, 문장 {file_sentence_count:,}개)"
        )

    if not all_sentences:
        raise ValueError("선택한 파일들에 학습할 문장이 없습니다.")

    if len(all_sentences) < 2:
        raise ValueError("전체 문장이 너무 적습니다. 최소 2개 이상 필요합니다.")

    log(f"선택한 파일 수: {len(paths):,}개")
    log(f"통합 문장 수: {len(all_sentences):,}개")

    return all_sentences, sentence_records


# =============================================================================
# 4. 형태소 분석과 토큰화
# =============================================================================

def build_tokenizer(
    user_words: list[str] | None,
    user_dict_path: str | None,
    alias_rules: list[tuple[str, list[str]]],
    token_pos_counts: dict[str, Counter],
    log: Callable[[str], None],
):
    """
    Kiwi 형태소 분석기를 만들고 tokenize 함수를 반환한다.

    kiwipiepy가 설치되지 않은 경우에는 None을 반환한다.
    호출한 쪽에서 공백 기반 대체 토크나이저를 사용할 수 있다.
    """

    try:
        from kiwipiepy import Kiwi
    except ImportError:
        return None

    # Kiwi 객체는 형태소 분석기의 본체다.
    kiwi = Kiwi()

    # 사용자가 GUI에 입력한 사용자 단어 목록을 복사한다.
    specs = list(user_words or [])

    # 이미 등록된 단어를 확인해 중복 등록을 줄인다.
    registered = {parse_user_word(spec)[0] for spec in specs}

    # 표현 통합 규칙의 대표어는 하나의 토큰으로 유지되어야 하므로
    # 자동으로 고유명사(NNP), 점수 5.0으로 등록한다.
    for canonical, _aliases in alias_rules:
        if canonical not in registered:
            specs.append(f"{canonical}:NNP:5.0")
            registered.add(canonical)

    # 사용자 단어와 자동 등록된 대표어를 Kiwi에 추가한다.
    for spec in specs:
        word, tag, score = parse_user_word(spec)

        # add_user_word가 False를 반환해도 이미 기본 사전에 있는 단어일 수 있으므로
        # 분석 전체를 중단하지 않고 로그만 남긴다.
        if not kiwi.add_user_word(word, tag, score):
            log(
                f"[알림] '{word}'({tag})는 이미 등록되었거나 "
                "기본 사전에 존재할 수 있습니다."
            )

    # 별도 사용자 사전 파일이 지정된 경우 불러온다.
    if user_dict_path:
        dictionary = Path(user_dict_path)

        if not dictionary.exists():
            raise FileNotFoundError(
                f"사용자 사전 파일을 찾을 수 없습니다: {dictionary}"
            )

        added = kiwi.load_user_dictionary(str(dictionary))
        log(f"사용자 사전 로드 완료: {dictionary.name} ({added}개 단어)")

    def tokenize(sentence: str) -> list[str]:
        """
        문장 한 개를 Word2Vec 학습용 단어 리스트로 변환한다.

        처리 순서
        ---------
        1. MZ세대 등의 표기를 대표어로 통합
        2. Kiwi 형태소 분석
        3. TRAIN_POS에 포함된 품사만 선택
        4. 단어별 품사 등장 횟수 기록
        """

        normalized = normalize_text(sentence, alias_rules)
        result: list[str] = []

        for token in kiwi.tokenize(normalized):
            if token.tag not in TRAIN_POS:
                continue

            result.append(token.form)

            # 나중에 결과를 명사 중심으로 제한할 때 사용한다.
            token_pos_counts[token.form][token.tag] += 1

        return result

    return tokenize


def build_simple_tokenizer(
    alias_rules: list[tuple[str, list[str]]],
    token_pos_counts: dict[str, Counter],
):
    """
    Kiwi가 설치되지 않았을 때 사용하는 대체 토크나이저를 만든다.

    형태소 분석을 하지 못하므로 정확도는 낮다.
    다만 프로그램이 완전히 실행되지 않는 것을 막기 위해 제공한다.
    """

    def tokenize(sentence: str) -> list[str]:
        # 먼저 표기 통합은 그대로 적용한다.
        normalized = normalize_text(sentence, alias_rules)

        # Kiwi가 없으면 조사와 대표어의 경계를 정확하게 찾기 어렵다.
        # 예를 들어 "MZ세대는"을 "MZ"로 통합하면 "MZ는"이 되는데,
        # 단순 정규표현식은 이를 "MZ는"이라는 한 단어로 잡을 수 있다.
        # 이를 줄이기 위해 대표어 앞뒤에 임시 공백을 넣어 토큰 경계를 만든다.
        for canonical, _aliases in alias_rules:
            normalized = re.sub(
                re.escape(canonical),
                f" {canonical} ",
                normalized,
                flags=re.IGNORECASE,
            )

        # 한글·영문·숫자·밑줄로 이루어진 연속 문자열을 단어처럼 추출한다.
        words = re.findall(r"[가-힣A-Za-z0-9_]+", normalized)

        # 실제 품사를 알 수 없으므로 임시로 NNP로 기록한다.
        for word in words:
            token_pos_counts[word]["NNP"] += 1

        return words

    return tokenize


def tokenize_sentences(
    sentences: list[str],
    tokenizer,
) -> list[list[str]]:
    """
    모든 문장에 토크나이저를 적용한다.

    gensim Word2Vec은 다음 형식의 데이터를 요구한다.

    [
        ["MZ", "소비", "문화"],
        ["4050", "소비", "시장"],
        ...
    ]
    """

    tokenized_sentences: list[list[str]] = []

    for sentence in sentences:
        tokens = tokenizer(sentence)

        # 의미 있는 토큰이 한 개도 없는 문장은 학습 목록에서 제외한다.
        if tokens:
            tokenized_sentences.append(tokens)

    return tokenized_sentences


# =============================================================================
# 5. Word2Vec 학습과 유사어 계산
# =============================================================================

def train_word2vec(
    tokenized_sentences: list[list[str]],
    vector_size: int,
    window: int,
    min_count: int,
    epochs: int,
    sg: int,
):
    """
    gensim Word2Vec 모델을 학습한다.

    주요 매개변수
    ------------
    vector_size
        단어 하나를 표현하는 벡터의 차원 수

    window
        기준 단어 앞뒤에서 문맥으로 볼 최대 단어 수

    min_count
        전체 말뭉치에서 이 횟수보다 적게 등장한 단어는 학습에서 제외

    epochs
        전체 학습 데이터를 반복해서 학습하는 횟수

    sg
        1 = Skip-gram
        0 = CBOW
    """

    try:
        from gensim.models import Word2Vec
    except ImportError as exc:
        raise ImportError(
            "gensim이 설치되어 있지 않습니다. "
            "터미널에서 'pip install gensim'을 실행하세요."
        ) from exc

    # workers가 너무 크면 작은 데이터에서 오히려 비효율적일 수 있어 최대 4개로 제한한다.
    worker_count = max(1, min(4, os.cpu_count() or 1))

    return Word2Vec(
        sentences=tokenized_sentences,
        vector_size=vector_size,
        window=window,
        min_count=min_count,
        epochs=epochs,
        sg=sg,
        workers=worker_count,
        # 같은 데이터와 설정에서 결과가 지나치게 달라지는 것을 줄이기 위한 시드다.
        seed=42,
    )


def get_similarity_results(
    model,
    query_words: list[str],
    top_n: int,
    excluded_words: set[str],
    result_mode: str,
    token_pos_counts: dict[str, Counter],
    log: Callable[[str], None],
) -> tuple[list[str], list[dict[str, object]]]:
    """
    기준어별 유사 단어를 계산하고 필터링한다.

    단순히 top_n개만 바로 가져오지 않는 이유
    -----------------------------------------
    상위 후보 중에는 제외어, 다른 기준어, 동사·형용사가 포함될 수 있다.
    따라서 top_n보다 훨씬 많은 후보를 먼저 가져온 뒤 조건에 맞는 단어만 남긴다.
    """

    vocab = model.wv.index_to_key
    log(f"학습된 어휘 수: {len(vocab):,}개")

    result_rows: list[dict[str, object]] = []
    found_targets: list[str] = []

    # 영문 대소문자를 무시하고 비교하기 위해 casefold 형태도 준비한다.
    excluded_folded = {word.casefold() for word in excluded_words}
    query_folded = {word.casefold() for word in query_words}

    for word in query_words:
        # min_count 때문에 기준어 자체가 학습 어휘에서 빠졌을 수 있다.
        if word not in model.wv:
            log(
                f"[건너뜀] '{word}'는 학습 어휘에 없습니다. "
                "표현 통합 규칙, 사용자 단어, 최소 등장 횟수를 확인하세요."
            )
            continue

        found_targets.append(word)

        # 자기 자신을 제외하면 최대 후보 수는 어휘 수 - 1이다.
        max_candidates = max(0, len(vocab) - 1)

        # 일단 top_n의 20배 또는 최소 200개를 후보로 가져온다.
        pool_size = min(max_candidates, max(200, top_n * 20))

        if pool_size:
            similar_items = model.wv.most_similar(word, topn=pool_size)
        else:
            similar_items = []

        selected: list[tuple[str, float]] = []

        for similar_word, score in similar_items:
            # 사용자가 지정한 결과 제외어는 건너뛴다.
            if similar_word.casefold() in excluded_folded:
                continue

            # 여러 기준어를 분석할 때 다른 기준어가 결과에 섞이는 것을 막는다.
            if similar_word.casefold() in query_folded:
                continue

            # 명사 중심 모드에서는 대표 품사가 명사·영문·숫자인 단어만 남긴다.
            if result_mode == "명사 중심":
                pos = dominant_pos(similar_word, token_pos_counts)

                if pos not in NOUN_RESULT_POS:
                    continue

            selected.append((similar_word, score))

            if len(selected) >= top_n:
                break

        # 위 후보 풀에서 원하는 개수를 채우지 못했다면 전체 후보를 다시 검색한다.
        if len(selected) < top_n and pool_size < max_candidates:
            all_items = model.wv.most_similar(word, topn=max_candidates)
            already_selected = {item[0] for item in selected}

            for similar_word, score in all_items:
                if similar_word in already_selected:
                    continue

                if similar_word.casefold() in excluded_folded:
                    continue

                if similar_word.casefold() in query_folded:
                    continue

                if result_mode == "명사 중심":
                    pos = dominant_pos(similar_word, token_pos_counts)

                    if pos not in NOUN_RESULT_POS:
                        continue

                selected.append((similar_word, score))

                if len(selected) >= top_n:
                    break

        log(f"'{word}' 유사어 {len(selected)}개 계산 완료")

        # Excel에 저장하기 쉬운 사전 형태로 결과를 만든다.
        for rank, (similar_word, score) in enumerate(selected, start=1):
            result_rows.append(
                {
                    "기준어": word,
                    "순위": rank,
                    "유사어": similar_word,
                    "유사도": round(float(score), 4),
                }
            )

    return found_targets, result_rows


def calculate_g_squared(
    pair_count: int,
    center_count: int,
    context_count: int,
    total_pairs: int,
) -> float:
    """단어 쌍의 2×2 분할표에서 로그우도비 G²를 계산한다."""

    observed = (
        pair_count,
        center_count - pair_count,
        context_count - pair_count,
        total_pairs - center_count - context_count + pair_count,
    )
    row_totals = (
        observed[0] + observed[1],
        observed[2] + observed[3],
    )
    column_totals = (
        observed[0] + observed[2],
        observed[1] + observed[3],
    )

    g_squared = 0.0

    for index, observed_count in enumerate(observed):
        if observed_count <= 0:
            continue

        row_index = 0 if index < 2 else 1
        column_index = 0 if index % 2 == 0 else 1
        expected_count = (
            row_totals[row_index]
            * column_totals[column_index]
            / total_pairs
        )

        if expected_count > 0:
            g_squared += observed_count * math.log(
                observed_count / expected_count
            )

    return 2.0 * g_squared


def get_pmi_results(
    tokenized_sentences: list[list[str]],
    query_words: list[str],
    top_n: int,
    window: int,
    min_count: int,
    excluded_words: set[str],
    result_mode: str,
    token_pos_counts: dict[str, Counter],
    log: Callable[[str], None],
) -> tuple[list[str], list[dict[str, object]]]:
    """
    기준어와 문맥 창 안에서 함께 등장한 단어의 양의 PMI를 계산한다.

    PMI(x, y) = log2(P(x, y) / (P(x) * P(y)))

    같은 문장 안에서도 기준 단어의 앞뒤 ``window`` 범위만 문맥으로 본다.
    드물게 등장한 단어가 과도하게 높은 점수를 받는 현상을 줄이기 위해
    ``min_count`` 미만인 단어는 계산에서 제외한다.
    """

    word_counts = Counter(
        token
        for sentence in tokenized_sentences
        for token in sentence
    )
    valid_words = {
        word
        for word, count in word_counts.items()
        if count >= min_count
    }
    query_word_set = set(query_words)

    pair_counts: Counter[tuple[str, str]] = Counter()
    center_counts: Counter[str] = Counter()
    context_counts: Counter[str] = Counter()
    total_pairs = 0

    for sentence in tokenized_sentences:
        sentence_length = len(sentence)

        for center_index, center_word in enumerate(sentence):
            if center_word not in valid_words:
                continue

            start = max(0, center_index - window)
            end = min(sentence_length, center_index + window + 1)

            for context_index in range(start, end):
                if context_index == center_index:
                    continue

                context_word = sentence[context_index]
                if context_word not in valid_words:
                    continue

                # 결과에 필요한 기준어 쌍만 저장해 큰 데이터의 메모리 사용을 줄인다.
                if center_word in query_word_set:
                    pair_counts[(center_word, context_word)] += 1

                center_counts[center_word] += 1
                context_counts[context_word] += 1
                total_pairs += 1

    result_rows: list[dict[str, object]] = []
    found_targets: list[str] = []
    excluded_folded = {word.casefold() for word in excluded_words}
    query_folded = {word.casefold() for word in query_words}

    if not total_pairs:
        return found_targets, result_rows

    for word in query_words:
        if word not in valid_words or not center_counts[word]:
            log(
                f"[건너뜀] '{word}'는 PMI 계산 어휘에 없습니다. "
                "표현 통합 규칙, 사용자 단어, 최소 등장 횟수를 확인하세요."
            )
            continue

        found_targets.append(word)
        candidates: list[tuple[str, float, float, int]] = []

        for (center_word, context_word), pair_count in pair_counts.items():
            if center_word != word:
                continue

            if context_word.casefold() in excluded_folded:
                continue

            if context_word.casefold() in query_folded:
                continue

            if result_mode == "명사 중심":
                pos = dominant_pos(context_word, token_pos_counts)

                if pos not in NOUN_RESULT_POS:
                    continue

            denominator = center_counts[word] * context_counts[context_word]
            score = math.log2((pair_count * total_pairs) / denominator)
            g_squared = calculate_g_squared(
                pair_count=pair_count,
                center_count=center_counts[word],
                context_count=context_counts[context_word],
                total_pairs=total_pairs,
            )

            # 관련성이 없는 음의 PMI는 유사어 결과에서 제외한다.
            if score <= 0:
                continue

            candidates.append(
                (context_word, score, g_squared, pair_count)
            )

        # PMI가 같으면 실제 동시 등장 횟수가 많은 단어를 먼저 보여준다.
        candidates.sort(key=lambda item: (item[1], item[3]), reverse=True)
        selected = candidates[:top_n]
        log(f"'{word}' PMI 연관어 {len(selected)}개 계산 완료")

        for rank, (similar_word, score, g_squared, _pair_count) in enumerate(
            selected,
            start=1,
        ):
            result_rows.append(
                {
                    "기준어": word,
                    "순위": rank,
                    "유사어": similar_word,
                    "유사도": round(float(score), 4),
                    "G²": round(float(g_squared), 4),
                }
            )

    return found_targets, result_rows


# =============================================================================
# 6. PCA 그래프 저장
# =============================================================================

def pca_2d(vectors):
    """
    고차원 Word2Vec 벡터를 2차원 좌표로 축소한다.

    그래프는 단어 관계를 직관적으로 보는 보조 수단이다.
    정확한 유사도 순위는 원래 고차원 벡터의 코사인 유사도로 계산된 Excel 결과를 봐야 한다.
    """

    try:
        import numpy as np
    except ImportError as exc:
        raise ImportError(
            "그래프 저장을 위해 numpy가 필요합니다. "
            "'pip install numpy'를 실행하세요."
        ) from exc

    vectors = np.asarray(vectors)

    # 각 차원의 평균을 빼서 데이터 중심을 원점으로 옮긴다.
    centered = vectors - vectors.mean(axis=0)

    # SVD로 분산이 가장 큰 축을 찾는다.
    _u, _s, vt = np.linalg.svd(centered, full_matrices=False)

    # 가장 중요한 두 축으로 투영하여 2차원 좌표를 만든다.
    return centered @ vt[:2].T


def setup_korean_font(log: Callable[[str], None]) -> None:
    """matplotlib 그래프에서 한글이 네모로 깨지는 것을 줄인다."""

    import matplotlib
    from matplotlib import font_manager

    # 운영체제별로 자주 설치되어 있는 한글 폰트 후보를 순서대로 검사한다.
    korean_fonts = ["Malgun Gothic", "AppleGothic", "NanumGothic"]
    available = {font.name for font in font_manager.fontManager.ttflist}

    for font_name in korean_fonts:
        if font_name in available:
            matplotlib.rcParams["font.family"] = font_name
            break
    else:
        log("[경고] 한글 폰트를 찾지 못해 그래프 글자가 깨질 수 있습니다.")

    # 음수 부호가 네모로 표시되는 문제를 막는다.
    matplotlib.rcParams["axes.unicode_minus"] = False


def plot_embeddings(
    model,
    path: str,
    target_words: list[str],
    result_rows: list[dict[str, object]],
    max_words: int,
    log: Callable[[str], None],
) -> None:
    """기준어와 유사어 벡터를 PCA 2D 그래프로 저장한다."""

    try:
        import matplotlib.pyplot as plt
    except ImportError as exc:
        raise ImportError(
            "그래프 저장을 위해 matplotlib이 필요합니다. "
            "'pip install matplotlib'을 실행하세요."
        ) from exc

    setup_korean_font(log)

    # 기준어를 먼저 넣고, 뒤에 유사어를 순서대로 추가한다.
    candidates = target_words + [
        str(row["유사어"])
        for row in result_rows
    ]

    vocab: list[str] = []
    seen: set[str] = set()

    for word in candidates:
        # 모델 어휘에 존재하고 아직 추가하지 않은 단어만 사용한다.
        if word in model.wv and word not in seen:
            vocab.append(word)
            seen.add(word)

        if len(vocab) >= max_words:
            break

    if len(vocab) < 2:
        log("[경고] 시각화할 단어가 2개 미만이라 그래프를 생략했습니다.")
        return

    vectors = [model.wv[word] for word in vocab]
    reduced = pca_2d(vectors)

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)

    plt.figure(figsize=(11, 8))
    plt.scatter(reduced[:, 0], reduced[:, 1], alpha=0.65)

    for word, (x, y) in zip(vocab, reduced):
        plt.annotate(
            word,
            (x, y),
            fontsize=9,
            xytext=(3, 3),
            textcoords="offset points",
        )

    plt.title("Word2Vec 임베딩 시각화 (PCA, 2D)")
    plt.tight_layout()
    plt.savefig(output, dpi=150)
    plt.close()

    log(f"그래프 저장 완료: {output}")


# =============================================================================
# 7. 출력 파일명과 Excel 저장
# =============================================================================

def make_safe_filename(text: str) -> str:
    """Windows 파일명에 사용할 수 없는 문자를 밑줄로 바꾼다."""

    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", text).strip()
    return cleaned or "Word2Vec"


def make_default_excel_path(
    targets: list[str],
    input_paths: list[str],
    analysis_mode: str,
) -> str:
    """
    Excel 저장 경로를 지정하지 않았을 때 자동 경로를 만든다.

    - 폴더: 첫 번째 입력 TXT 파일과 같은 폴더
    - 파일명: 기준어_분석모드_결과_날짜시간.xlsx
    """

    keyword = make_safe_filename(targets[0]) if targets else analysis_mode
    mode_slug = "pmi" if analysis_mode == "PMI" else "word2vec"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{keyword}_{mode_slug}_결과_{timestamp}.xlsx"

    first_input_folder = Path(input_paths[0]).expanduser().resolve().parent
    return str(first_input_folder / filename)


def save_excel_report(
    output_path: str,
    sentence_records: list[dict[str, object]],
    input_file_names: list[str],
    requested_targets: list[str],
    normalized_targets: list[str],
    found_targets: list[str],
    result_rows: list[dict[str, object]],
    source: str,
    settings_text: str,
    alias_rules_text: str,
    excluded_words_text: str,
    result_mode: str,
    analysis_mode: str,
) -> None:
    """
    분석 결과를 Excel 파일로 저장한다.

    생성 시트
    ---------
    1. 분석 결과
       분석 조건과 유사어 순위 저장

    2. 수집 기사 제목
       각 원문 문장과 원본 파일명, 파일 내 줄 번호 저장
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise ImportError(
            "엑셀 저장을 위해 openpyxl이 필요합니다. "
            "'pip install openpyxl'을 실행하세요."
        ) from exc

    # 새 Excel 통합 문서를 생성한다.
    workbook = Workbook()

    # 기본으로 생성된 첫 시트를 분석 결과 시트로 사용한다.
    result_sheet = workbook.active
    result_sheet.title = f"{analysis_mode} 결과"

    # 두 번째 시트는 원문 목록을 저장한다.
    article_sheet = workbook.create_sheet("수집 기사 제목")

    # -------------------------------------------------------------------------
    # Excel 스타일 설정
    # -------------------------------------------------------------------------
    dark_blue = "1F4E78"
    medium_blue = "5B9BD5"
    light_blue = "D9EAF7"
    very_light_blue = "EDF4FA"
    white = "FFFFFF"
    gray = "666666"

    thin_side = Side(style="thin", color="B7C9D6")
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    title_font = Font(size=15, bold=True, color=white)
    label_font = Font(bold=True, color="1F1F1F")
    header_font = Font(bold=True, color="1F1F1F")
    note_font = Font(size=9, color=gray)

    title_fill = PatternFill("solid", fgColor=dark_blue)
    label_fill = PatternFill("solid", fgColor=light_blue)
    header_fill = PatternFill("solid", fgColor=medium_blue)
    alternate_fill = PatternFill("solid", fgColor=very_light_blue)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrapped_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # -------------------------------------------------------------------------
    # 첫 번째 시트: 선택한 분석 모드의 연관어
    # -------------------------------------------------------------------------
    result_column_count = 5 if analysis_mode == "PMI" else 4
    result_last_column = "E" if analysis_mode == "PMI" else "D"

    result_sheet.merge_cells(f"A1:{result_last_column}1")
    result_sheet["A1"] = f"뉴스 텍스트 기반 {analysis_mode} 연관어 분석"
    result_sheet["A1"].font = title_font
    result_sheet["A1"].fill = title_fill
    result_sheet["A1"].alignment = center
    result_sheet.row_dimensions[1].height = 28

    # 입력 기준어와 실제 통합 후 기준어를 모두 기록한다.
    requested_text = ", ".join(requested_targets)
    normalized_text = ", ".join(normalized_targets)
    found_text = ", ".join(found_targets) if found_targets else "없음"

    metadata = [
        ("입력 기준 키워드", requested_text),
        ("통합 후 기준 키워드", normalized_text),
        ("분석 성공 키워드", found_text),
        ("출처", source),
        ("입력 파일 수", len(input_file_names)),
        ("입력 파일 목록", ", ".join(input_file_names)),
        ("전체 문장 수", len(sentence_records)),
        ("학습 설정", settings_text),
        ("표현 통합 규칙", alias_rules_text or "없음"),
        ("결과 제외어", excluded_words_text or "없음"),
        ("결과 품사", result_mode),
        ("결과 생성일", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    # 메타데이터는 3행부터 작성한다.
    for row_index, (label, value) in enumerate(metadata, start=3):
        label_cell = result_sheet.cell(row=row_index, column=1, value=label)
        value_cell = result_sheet.cell(row=row_index, column=2, value=value)

        # 값 영역은 B열부터 결과 표의 마지막 열까지 합친다.
        result_sheet.merge_cells(
            start_row=row_index,
            start_column=2,
            end_row=row_index,
            end_column=result_column_count,
        )

        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.alignment = center
        label_cell.border = thin_border

        value_cell.alignment = wrapped_left
        value_cell.border = thin_border

        # 병합된 셀의 테두리가 끊겨 보이지 않도록 전체에 테두리를 지정한다.
        for column_index in range(2, result_column_count + 1):
            result_sheet.cell(row=row_index, column=column_index).border = thin_border

    # 메타데이터 아래에 한 줄 간격을 두고 결과 표 머리글을 만든다.
    header_row = 4 + len(metadata)
    if analysis_mode == "PMI":
        headers = [
            "분석 기준 키워드",
            "순위",
            "PMI 연관어",
            "PMI 점수",
            "G²",
        ]
    else:
        headers = [
            "분석 기준 키워드",
            "순위",
            "Word2Vec 연관어",
            "유사도",
        ]

    for column_index, header in enumerate(headers, start=1):
        cell = result_sheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # 실제 유사어 결과를 한 행씩 작성한다.
    for row_index, item in enumerate(result_rows, start=header_row + 1):
        values = [
            item["기준어"],
            item["순위"],
            item["유사어"],
            item["유사도"],
        ]

        if analysis_mode == "PMI":
            values.append(item["G²"])

        for column_index, value in enumerate(values, start=1):
            cell = result_sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )
            cell.alignment = center
            cell.border = thin_border

            # 읽기 쉽도록 짝수 행에 연한 배경색을 적용한다.
            if row_index % 2 == 0:
                cell.fill = alternate_fill

            # 유사도·PMI·G² 점수는 소수점 네 자리까지 표시한다.
            if column_index >= 4:
                cell.number_format = "0.0000"

    # 기준어를 찾지 못한 경우 빈 표 대신 안내문을 표시한다.
    if not result_rows:
        result_sheet.merge_cells(
            start_row=header_row + 1,
            start_column=1,
            end_row=header_row + 2,
            end_column=result_column_count,
        )

        no_result = result_sheet.cell(
            row=header_row + 1,
            column=1,
            value=(
                "학습 어휘에서 기준 키워드를 찾지 못해 "
                "유사어 결과가 없습니다."
            ),
        )
        no_result.alignment = center
        no_result.font = note_font

    # 열 너비를 보기 좋게 조정한다.
    result_sheet.column_dimensions["A"].width = 22
    result_sheet.column_dimensions["B"].width = 10
    result_sheet.column_dimensions["C"].width = 26
    result_sheet.column_dimensions["D"].width = 14
    if analysis_mode == "PMI":
        result_sheet.column_dimensions["E"].width = 14

    # 결과 표 머리글 아래부터 스크롤되도록 틀을 고정한다.
    result_sheet.freeze_panes = f"A{header_row + 1}"

    # 자동 필터를 적용한다.
    result_last_row = max(header_row, header_row + len(result_rows))
    result_sheet.auto_filter.ref = (
        f"A{header_row}:{result_last_column}{result_last_row}"
    )

    # -------------------------------------------------------------------------
    # 두 번째 시트: 원본 문장 목록
    # -------------------------------------------------------------------------
    article_headers = ["순번", "원본 파일명", "파일 내 줄 번호", "기사 제목 또는 입력 문장"]

    for column_index, header in enumerate(article_headers, start=1):
        cell = article_sheet.cell(row=1, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for row_index, record in enumerate(sentence_records, start=2):
        values = [
            row_index - 1,
            record["파일명"],
            record["줄번호"],
            record["문장"],
        ]

        for column_index, value in enumerate(values, start=1):
            cell = article_sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )
            cell.border = thin_border

            # 문장 열만 왼쪽 정렬과 줄바꿈을 사용한다.
            if column_index == 4:
                cell.alignment = wrapped_left
            else:
                cell.alignment = center

            if row_index % 2 == 0:
                cell.fill = alternate_fill

    article_sheet.column_dimensions["A"].width = 10
    article_sheet.column_dimensions["B"].width = 28
    article_sheet.column_dimensions["C"].width = 14
    article_sheet.column_dimensions["D"].width = 100
    article_sheet.freeze_panes = "A2"
    article_sheet.auto_filter.ref = f"A1:D{len(sentence_records) + 1}"

    # 저장 폴더가 없으면 자동 생성한다.
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    # 최종 Excel 파일 저장
    workbook.save(output)


# =============================================================================
# 8. GUI 클래스
# =============================================================================

class Word2VecGUI:
    """프로그램의 화면과 사용자 입력, 분석 실행을 관리하는 메인 GUI 클래스다."""

    def __init__(self, root: tk.Tk):
        # tkinter 최상위 창 객체를 저장한다.
        self.root = root

        # 창 제목과 기본 크기를 설정한다.
        self.root.title("한국어 Word2Vec·PMI 문맥 연관어 분석기")
        self.root.geometry("1000x900")
        self.root.minsize(900, 760)

        # 백그라운드 분석 스레드가 GUI에 메시지를 전달하는 큐다.
        self.message_queue: queue.Queue[tuple[str, object]] = queue.Queue()

        # 사용자가 선택한 실제 파일 경로 목록이다.
        # 화면 Entry에는 보기 편한 문자열만 표시하고 실제 처리는 이 리스트를 사용한다.
        self.input_paths: list[str] = []

        # 마지막으로 저장한 Excel 경로다.
        # 분석 완료 후 '결과 폴더 열기' 버튼에서 사용한다.
        self.last_output_path: Path | None = None

        # 중복 실행을 막는 상태 변수다.
        self.running = False

        # GUI 변수, 스타일, 위젯을 순서대로 만든다.
        self._create_variables()
        self._configure_style()
        self._build_ui()
        self._toggle_optional_outputs()

        # 100ms마다 메시지 큐를 확인한다.
        self.root.after(100, self._process_queue)

    # -------------------------------------------------------------------------
    # GUI 상태 변수
    # -------------------------------------------------------------------------
    def _create_variables(self) -> None:
        """Entry, Checkbutton, Combobox와 연결할 tkinter 변수를 생성한다."""

        # 선택한 파일 목록을 화면에 표시하는 문자열이다.
        self.input_display_var = tk.StringVar()

        # 분석 기준 키워드 입력란이다.
        self.keywords_var = tk.StringVar()

        # Excel 메타데이터에 기록할 출처 설명이다.
        self.source_var = tk.StringVar(value="입력 TXT 파일")

        # Kiwi에 추가 등록할 사용자 단어 입력란이다.
        self.user_words_var = tk.StringVar()

        # Kiwi 사용자 사전 파일 경로다.
        self.user_dict_var = tk.StringVar()

        # 같은 의미의 다른 표기를 대표어로 통합하는 규칙이다.
        self.alias_rules_var = tk.StringVar(value=DEFAULT_ALIAS_RULES)

        # 유사어 결과에서 제외할 단어들이다.
        self.excluded_words_var = tk.StringVar(value=DEFAULT_EXCLUDED_WORDS)

        # 분석 모드와 학습 및 출력 설정 기본값이다.
        self.analysis_mode_var = tk.StringVar(value="Word2Vec")
        self.recommendation_var = tk.StringVar(
            value="분석을 시작하면 데이터에 맞는 모드를 추천합니다."
        )
        self.top_n_var = tk.StringVar(value="30")
        self.vector_size_var = tk.StringVar(value="200")
        self.window_var = tk.StringVar(value="5")
        self.min_count_var = tk.StringVar(value="2")
        self.epochs_var = tk.StringVar(value="30")
        self.method_var = tk.StringVar(value="Skip-gram")
        self.result_mode_var = tk.StringVar(value="명사 중심")

        # 결과 Excel 저장 경로다. 비워 두면 자동 생성한다.
        self.excel_var = tk.StringVar()

        # Word2Vec 모델 저장 여부와 경로다.
        self.save_model_var = tk.BooleanVar(value=False)
        self.model_var = tk.StringVar()

        # PCA 그래프 저장 여부와 경로다.
        self.save_plot_var = tk.BooleanVar(value=False)
        self.plot_var = tk.StringVar()
        self.plot_max_words_var = tk.StringVar(value="50")

        # 화면 오른쪽 아래에 표시할 현재 상태 문자열이다.
        self.status_var = tk.StringVar(value="대기 중")

    # -------------------------------------------------------------------------
    # GUI 스타일
    # -------------------------------------------------------------------------
    def _configure_style(self) -> None:
        """ttk 위젯의 글꼴과 일부 스타일을 설정한다."""

        style = ttk.Style()

        # Windows에서는 vista 테마가 비교적 자연스럽다.
        # 지원하지 않는 환경에서는 기본 테마를 그대로 사용한다.
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass

        style.configure("Title.TLabel", font=("맑은 고딕", 17, "bold"))
        style.configure(
            "Section.TLabelframe.Label",
            font=("맑은 고딕", 10, "bold"),
        )
        style.configure("Run.TButton", font=("맑은 고딕", 10, "bold"), padding=7)
        style.configure("Hint.TLabel", foreground="#666666")

    # -------------------------------------------------------------------------
    # 전체 화면 구성
    # -------------------------------------------------------------------------
    def _build_ui(self) -> None:
        """창 안에 탭, 버튼, 진행바, 로그창을 배치한다."""

        outer = ttk.Frame(self.root, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(
            outer,
            text="한국어 Word2Vec·PMI 문맥 연관어 분석기",
            style="Title.TLabel",
        ).pack(anchor="w")

        ttk.Label(
            outer,
            text=(
                "여러 일반 뉴스 파일 전체를 분석해 "
                "기준어와 문맥상 연관된 단어를 찾습니다."
            ),
            style="Hint.TLabel",
        ).pack(anchor="w", pady=(2, 12))

        # 설정을 세 개 탭으로 구분한다.
        notebook = ttk.Notebook(outer)
        notebook.pack(fill="x")

        basic_tab = ttk.Frame(notebook, padding=12)
        advanced_tab = ttk.Frame(notebook, padding=12)
        output_tab = ttk.Frame(notebook, padding=12)

        notebook.add(basic_tab, text="기본 설정")
        notebook.add(advanced_tab, text="학습 설정")
        notebook.add(output_tab, text="저장 설정")

        self._build_basic_tab(basic_tab)
        self._build_advanced_tab(advanced_tab)
        self._build_output_tab(output_tab)

        # 분석 실행 및 보조 버튼 영역
        controls = ttk.Frame(outer)
        controls.pack(fill="x", pady=(12, 8))

        self.run_button = ttk.Button(
            controls,
            text="분석 시작",
            command=self.start_analysis,
            style="Run.TButton",
        )
        self.run_button.pack(side="left")

        self.open_folder_button = ttk.Button(
            controls,
            text="결과 폴더 열기",
            command=self.open_result_folder,
            state="disabled",
        )
        self.open_folder_button.pack(side="left", padx=(8, 0))

        ttk.Button(
            controls,
            text="로그 지우기",
            command=self.clear_log,
        ).pack(side="right")

        # 진행바와 현재 상태 문자열
        progress_frame = ttk.Frame(outer)
        progress_frame.pack(fill="x", pady=(0, 8))

        self.progress = ttk.Progressbar(progress_frame, mode="indeterminate")
        self.progress.pack(side="left", fill="x", expand=True)

        ttk.Label(
            progress_frame,
            textvariable=self.status_var,
            width=18,
            anchor="e",
        ).pack(side="right", padx=(10, 0))

        # 분석 진행 내용을 출력하는 로그 창
        log_frame = ttk.LabelFrame(
            outer,
            text="진행 로그",
            padding=8,
            style="Section.TLabelframe",
        )
        log_frame.pack(fill="both", expand=True)

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            height=16,
            wrap="word",
            state="disabled",
            font=("Consolas", 9),
        )
        self.log_text.pack(fill="both", expand=True)

    # -------------------------------------------------------------------------
    # 기본 설정 탭
    # -------------------------------------------------------------------------
    def _build_basic_tab(self, parent: ttk.Frame) -> None:
        """입력 파일, 기준어, 표현 통합, 제외어 설정 화면을 만든다."""

        parent.columnconfigure(1, weight=1)

        # 여러 TXT 파일 선택
        ttk.Label(parent, text="학습 TXT 파일들").grid(
            row=0,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(
            parent,
            textvariable=self.input_display_var,
            state="readonly",
        ).grid(row=0, column=1, sticky="ew", padx=8, pady=6)

        ttk.Button(
            parent,
            text="여러 파일 선택",
            command=self.browse_inputs,
        ).grid(row=0, column=2, pady=6)

        ttk.Label(
            parent,
            text="Ctrl 또는 Shift를 사용해 여러 TXT 파일을 동시에 선택할 수 있습니다.",
            style="Hint.TLabel",
        ).grid(row=1, column=1, columnspan=2, sticky="w", padx=(8, 0))

        # 기준어 입력
        ttk.Label(parent, text="분석 기준 키워드").grid(
            row=2,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(parent, textvariable=self.keywords_var).grid(
            row=2,
            column=1,
            columnspan=2,
            sticky="ew",
            padx=(8, 0),
            pady=6,
        )

        ttk.Label(
            parent,
            text="여러 개는 쉼표로 구분합니다. 예: MZ세대, 청년, 4050",
            style="Hint.TLabel",
        ).grid(row=3, column=1, columnspan=2, sticky="w", padx=(8, 0))

        # 출처 표시
        ttk.Label(parent, text="Excel 출처 표시").grid(
            row=4,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(parent, textvariable=self.source_var).grid(
            row=4,
            column=1,
            columnspan=2,
            sticky="ew",
            padx=(8, 0),
            pady=6,
        )

        ttk.Separator(parent).grid(
            row=5,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=10,
        )

        # 사용자 단어
        ttk.Label(parent, text="추가 사용자 단어").grid(
            row=6,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(parent, textvariable=self.user_words_var).grid(
            row=6,
            column=1,
            columnspan=2,
            sticky="ew",
            padx=(8, 0),
            pady=6,
        )

        ttk.Label(
            parent,
            text="선택 사항. 쉼표 구분: 갓생:NNP:5.0, 영끌:NNP:5.0",
            style="Hint.TLabel",
        ).grid(row=7, column=1, columnspan=2, sticky="w", padx=(8, 0))

        # 사용자 사전 파일
        ttk.Label(parent, text="사용자 사전 파일").grid(
            row=8,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(parent, textvariable=self.user_dict_var).grid(
            row=8,
            column=1,
            sticky="ew",
            padx=8,
            pady=6,
        )

        ttk.Button(
            parent,
            text="찾아보기",
            command=self.browse_user_dict,
        ).grid(row=8, column=2, pady=6)

        ttk.Separator(parent).grid(
            row=9,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=10,
        )

        # 표현 통합 규칙
        ttk.Label(parent, text="표현 통합 규칙").grid(
            row=10,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(parent, textvariable=self.alias_rules_var).grid(
            row=10,
            column=1,
            columnspan=2,
            sticky="ew",
            padx=(8, 0),
            pady=6,
        )

        ttk.Label(
            parent,
            text=(
                "대표어=표현1|표현2 형식, 규칙 사이는 세미콜론. "
                "예: MZ=MZ세대|MZ 세대"
            ),
            style="Hint.TLabel",
        ).grid(row=11, column=1, columnspan=2, sticky="w", padx=(8, 0))

        # 결과 제외어
        ttk.Label(parent, text="결과 제외어").grid(
            row=12,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(parent, textvariable=self.excluded_words_var).grid(
            row=12,
            column=1,
            columnspan=2,
            sticky="ew",
            padx=(8, 0),
            pady=6,
        )

        ttk.Label(
            parent,
            text=(
                "쉼표 구분. 기준어와 함께 자주 붙지만 비교 대상이 아닌 단어를 "
                "결과에서 제외합니다."
            ),
            style="Hint.TLabel",
        ).grid(row=13, column=1, columnspan=2, sticky="w", padx=(8, 0))

    # -------------------------------------------------------------------------
    # 학습 설정 탭
    # -------------------------------------------------------------------------
    def _build_advanced_tab(self, parent: ttk.Frame) -> None:
        """Word2Vec의 주요 학습 매개변수 설정 화면을 만든다."""

        for column_index in range(4):
            parent.columnconfigure(column_index, weight=1)

        fields = [
            ("유사어 개수", self.top_n_var, "기준어마다 출력할 유사어 수"),
            ("벡터 차원", self.vector_size_var, "단어 벡터의 차원 수"),
            ("문맥 범위", self.window_var, "앞뒤 몇 단어까지 문맥으로 볼지"),
            ("최소 등장 횟수", self.min_count_var, "이 횟수 미만 단어는 제외"),
            ("학습 반복 횟수", self.epochs_var, "전체 문장을 반복 학습하는 횟수"),
            (
                "그래프 최대 단어",
                self.plot_max_words_var,
                "PCA 그래프에 표시할 최대 단어 수",
            ),
        ]

        # 두 항목씩 한 줄에 배치한다.
        for index, (label, variable, hint) in enumerate(fields):
            row = (index // 2) * 2
            column = (index % 2) * 2

            ttk.Label(parent, text=label).grid(
                row=row,
                column=column,
                sticky="w",
                padx=(0, 8),
                pady=7,
            )

            ttk.Spinbox(
                parent,
                from_=1,
                to=10000,
                textvariable=variable,
                width=12,
            ).grid(row=row, column=column + 1, sticky="w", pady=7)

            ttk.Label(
                parent,
                text=hint,
                style="Hint.TLabel",
            ).grid(
                row=row + 1,
                column=column,
                columnspan=2,
                sticky="w",
                pady=(0, 6),
            )

        # 학습 방식 선택
        method_row = 6

        ttk.Label(parent, text="학습 방식").grid(
            row=method_row,
            column=0,
            sticky="w",
            pady=7,
        )

        ttk.Combobox(
            parent,
            textvariable=self.method_var,
            values=["Skip-gram", "CBOW"],
            state="readonly",
            width=16,
        ).grid(row=method_row, column=1, sticky="w", pady=7)

        ttk.Label(
            parent,
            text="Skip-gram은 희귀 단어에 비교적 강하고, CBOW는 더 빠릅니다.",
            style="Hint.TLabel",
        ).grid(row=method_row + 1, column=0, columnspan=4, sticky="w")

        # 결과 품사 선택
        ttk.Label(parent, text="결과 품사").grid(
            row=method_row + 2,
            column=0,
            sticky="w",
            pady=7,
        )

        ttk.Combobox(
            parent,
            textvariable=self.result_mode_var,
            values=["명사 중심", "내용어 전체"],
            state="readonly",
            width=16,
        ).grid(row=method_row + 2, column=1, sticky="w", pady=7)

        ttk.Label(
            parent,
            text=(
                "학습은 명사·영문·숫자·동사·형용사·부사를 모두 사용하고, "
                "출력 단계에서만 품사를 제한합니다."
            ),
            style="Hint.TLabel",
        ).grid(row=method_row + 3, column=0, columnspan=4, sticky="w")

        # Word2Vec과 PMI 분석 모드 선택
        ttk.Label(parent, text="분석 모드").grid(
            row=method_row + 4,
            column=0,
            sticky="w",
            pady=7,
        )

        ttk.Combobox(
            parent,
            textvariable=self.analysis_mode_var,
            values=["Word2Vec", "PMI"],
            state="readonly",
            width=16,
        ).grid(row=method_row + 4, column=1, sticky="w", pady=7)

        self.analysis_mode_var.trace_add(
            "write",
            lambda *_args: self._toggle_optional_outputs(),
        )

        ttk.Label(
            parent,
            textvariable=self.recommendation_var,
            style="Hint.TLabel",
        ).grid(row=method_row + 5, column=0, columnspan=4, sticky="w")

    # -------------------------------------------------------------------------
    # 저장 설정 탭
    # -------------------------------------------------------------------------
    def _build_output_tab(self, parent: ttk.Frame) -> None:
        """Excel, 모델, 그래프 저장 설정 화면을 만든다."""

        parent.columnconfigure(1, weight=1)

        # Excel 저장 위치
        ttk.Label(parent, text="결과 Excel 파일").grid(
            row=0,
            column=0,
            sticky="w",
            pady=6,
        )

        ttk.Entry(parent, textvariable=self.excel_var).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=8,
            pady=6,
        )

        ttk.Button(
            parent,
            text="저장 위치",
            command=self.browse_excel,
        ).grid(row=0, column=2, pady=6)

        ttk.Label(
            parent,
            text="비워 두면 첫 번째 입력 TXT 파일과 같은 폴더에 자동 저장합니다.",
            style="Hint.TLabel",
        ).grid(row=1, column=1, columnspan=2, sticky="w", padx=(8, 0))

        ttk.Separator(parent).grid(
            row=2,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=12,
        )

        # 모델 저장 선택
        self.save_model_check = ttk.Checkbutton(
            parent,
            text="학습 모델도 저장",
            variable=self.save_model_var,
            command=self._toggle_optional_outputs,
        )
        self.save_model_check.grid(row=3, column=0, sticky="w", pady=6)

        self.model_entry = ttk.Entry(parent, textvariable=self.model_var)
        self.model_entry.grid(row=3, column=1, sticky="ew", padx=8, pady=6)

        self.model_button = ttk.Button(
            parent,
            text="저장 위치",
            command=self.browse_model,
        )
        self.model_button.grid(row=3, column=2, pady=6)

        # 그래프 저장 선택
        self.save_plot_check = ttk.Checkbutton(
            parent,
            text="PCA 그래프도 저장",
            variable=self.save_plot_var,
            command=self._toggle_optional_outputs,
        )
        self.save_plot_check.grid(row=4, column=0, sticky="w", pady=6)

        self.plot_entry = ttk.Entry(parent, textvariable=self.plot_var)
        self.plot_entry.grid(row=4, column=1, sticky="ew", padx=8, pady=6)

        self.plot_button = ttk.Button(
            parent,
            text="저장 위치",
            command=self.browse_plot,
        )
        self.plot_button.grid(row=4, column=2, pady=6)

    # -------------------------------------------------------------------------
    # 파일 선택 창
    # -------------------------------------------------------------------------
    def browse_inputs(self) -> None:
        """여러 TXT 파일을 동시에 선택한다."""

        selected = filedialog.askopenfilenames(
            title="분석용 TXT 파일 여러 개 선택",
            filetypes=[
                ("텍스트 파일", "*.txt"),
                ("모든 파일", "*.*"),
            ],
        )

        if not selected:
            return

        self.input_paths = list(selected)

        # Entry에는 전체 경로 대신 파일명만 표시한다.
        display_names = [Path(path).name for path in self.input_paths]
        self.input_display_var.set("; ".join(display_names))

        # 기본 출처 문자열이면 선택한 파일 수를 반영한다.
        if self.source_var.get().strip() in {"", "입력 TXT 파일"}:
            self.source_var.set(f"입력 TXT 파일 {len(self.input_paths)}개")

    def browse_user_dict(self) -> None:
        """Kiwi 사용자 사전 파일을 선택한다."""

        selected = filedialog.askopenfilename(
            title="Kiwi 사용자 사전 선택",
            filetypes=[
                ("텍스트 파일", "*.txt"),
                ("모든 파일", "*.*"),
            ],
        )

        if selected:
            self.user_dict_var.set(selected)

    def browse_excel(self) -> None:
        """결과 Excel 저장 위치를 선택한다."""

        selected = filedialog.asksaveasfilename(
            title="결과 Excel 저장 위치",
            defaultextension=".xlsx",
            filetypes=[("Excel 통합 문서", "*.xlsx")],
        )

        if selected:
            self.excel_var.set(selected)

    def browse_model(self) -> None:
        """Word2Vec 모델 저장 위치를 선택한다."""

        selected = filedialog.asksaveasfilename(
            title="Word2Vec 모델 저장 위치",
            defaultextension=".model",
            filetypes=[
                ("Word2Vec 모델", "*.model"),
                ("모든 파일", "*.*"),
            ],
        )

        if selected:
            self.model_var.set(selected)

    def browse_plot(self) -> None:
        """PCA 그래프 PNG 저장 위치를 선택한다."""

        selected = filedialog.asksaveasfilename(
            title="PCA 그래프 저장 위치",
            defaultextension=".png",
            filetypes=[("PNG 이미지", "*.png")],
        )

        if selected:
            self.plot_var.set(selected)

    def _toggle_optional_outputs(self) -> None:
        """체크 여부에 따라 모델·그래프 저장 경로 입력란을 활성화한다."""

        is_word2vec = self.analysis_mode_var.get() == "Word2Vec"

        if not is_word2vec:
            self.save_model_var.set(False)
            self.save_plot_var.set(False)

        check_state = "normal" if is_word2vec else "disabled"
        model_state = (
            "normal"
            if is_word2vec and self.save_model_var.get()
            else "disabled"
        )
        plot_state = (
            "normal"
            if is_word2vec and self.save_plot_var.get()
            else "disabled"
        )

        self.save_model_check.configure(state=check_state)
        self.save_plot_check.configure(state=check_state)
        self.model_entry.configure(state=model_state)
        self.model_button.configure(state=model_state)
        self.plot_entry.configure(state=plot_state)
        self.plot_button.configure(state=plot_state)

    # -------------------------------------------------------------------------
    # 입력값 검증
    # -------------------------------------------------------------------------
    @staticmethod
    def _parse_positive_int(
        value: str,
        label: str,
        minimum: int = 1,
    ) -> int:
        """문자열 입력값이 minimum 이상의 정수인지 확인한다."""

        try:
            number = int(value.strip())
        except ValueError as exc:
            raise ValueError(f"'{label}' 값은 정수여야 합니다.") from exc

        if number < minimum:
            raise ValueError(f"'{label}' 값은 {minimum} 이상이어야 합니다.")

        return number

    def _collect_settings(self) -> dict[str, object]:
        """
        GUI에 입력된 값을 읽고 검증한 뒤 하나의 설정 사전으로 만든다.

        분석 스레드는 tkinter 변수에 직접 접근하지 않고 이 사전만 사용한다.
        tkinter 위젯은 기본적으로 메인 스레드에서만 접근하는 것이 안전하기 때문이다.
        """

        if not self.input_paths:
            raise ValueError("학습 TXT 파일을 한 개 이상 선택하세요.")

        # 쉼표로 구분된 기준어를 리스트로 만든다.
        requested_query_words = [
            word.strip()
            for word in self.keywords_var.get().split(",")
            if word.strip()
        ]

        if not requested_query_words:
            raise ValueError("분석 기준 키워드를 입력하세요.")

        # 표현 통합 규칙을 파싱한다.
        alias_rules_text = self.alias_rules_var.get().strip()
        alias_rules = parse_alias_rules(alias_rules_text)

        # 사용자가 MZ세대를 입력했더라도 규칙에 따라 MZ로 통합한다.
        normalized_query_words: list[str] = []

        for word in requested_query_words:
            normalized = normalize_term(word, alias_rules)

            # 동일한 대표어가 중복되지 않도록 한다.
            if normalized not in normalized_query_words:
                normalized_query_words.append(normalized)

        # 결과 제외어를 리스트와 집합으로 만든다.
        excluded_words_raw = [
            word.strip()
            for word in self.excluded_words_var.get().split(",")
            if word.strip()
        ]

        excluded_words = {
            normalize_term(word, alias_rules)
            for word in excluded_words_raw
        }

        # 사용자 단어도 쉼표로 구분한다.
        user_words = [
            word.strip()
            for word in self.user_words_var.get().split(",")
            if word.strip()
        ]

        # 숫자 설정을 검증한다.
        top_n = self._parse_positive_int(self.top_n_var.get(), "유사어 개수")
        vector_size = self._parse_positive_int(
            self.vector_size_var.get(),
            "벡터 차원",
            2,
        )
        window = self._parse_positive_int(self.window_var.get(), "문맥 범위")
        min_count = self._parse_positive_int(
            self.min_count_var.get(),
            "최소 등장 횟수",
        )
        epochs = self._parse_positive_int(
            self.epochs_var.get(),
            "학습 반복 횟수",
        )
        plot_max_words = self._parse_positive_int(
            self.plot_max_words_var.get(),
            "그래프 최대 단어",
            2,
        )

        # gensim에서는 sg=1이 Skip-gram, sg=0이 CBOW다.
        sg = 1 if self.method_var.get() == "Skip-gram" else 0
        analysis_mode = self.analysis_mode_var.get()

        # Excel 경로가 비어 있으면 자동 경로를 만든다.
        excel_path = self.excel_var.get().strip().strip('"')

        if not excel_path:
            excel_path = make_default_excel_path(
                normalized_query_words,
                self.input_paths,
                analysis_mode,
            )

        # 모델/그래프 경로가 비어 있으면 Excel 파일명과 같은 이름을 사용한다.
        model_path = self.model_var.get().strip().strip('"')
        plot_path = self.plot_var.get().strip().strip('"')
        base_output = Path(excel_path)

        if self.save_model_var.get() and not model_path:
            model_path = str(base_output.with_suffix(".model"))

        if self.save_plot_var.get() and not plot_path:
            plot_path = str(base_output.with_suffix(".png"))

        source = self.source_var.get().strip()

        if not source:
            source = f"입력 TXT 파일 {len(self.input_paths)}개"

        return {
            "input_paths": list(self.input_paths),
            "input_file_names": [Path(path).name for path in self.input_paths],
            "requested_query_words": requested_query_words,
            "query_words": normalized_query_words,
            "source": source,
            "user_words": user_words,
            "user_dict_path": self.user_dict_var.get().strip().strip('"') or None,
            "alias_rules": alias_rules,
            "alias_rules_text": alias_rules_text,
            "excluded_words": excluded_words,
            "excluded_words_text": ", ".join(excluded_words_raw),
            "result_mode": self.result_mode_var.get(),
            "analysis_mode": analysis_mode,
            "top_n": top_n,
            "vector_size": vector_size,
            "window": window,
            "min_count": min_count,
            "epochs": epochs,
            "sg": sg,
            "plot_max_words": plot_max_words,
            "excel_path": excel_path,
            "save_model": self.save_model_var.get(),
            "model_path": model_path,
            "save_plot": self.save_plot_var.get(),
            "plot_path": plot_path,
        }

    # -------------------------------------------------------------------------
    # 분석 시작
    # -------------------------------------------------------------------------
    def start_analysis(self) -> None:
        """입력값을 검증하고 백그라운드 분석 스레드를 시작한다."""

        # 이미 분석 중이면 다시 실행하지 않는다.
        if self.running:
            return

        try:
            settings = self._collect_settings()
        except Exception as exc:
            messagebox.showerror("입력 오류", str(exc))
            return

        self.clear_log()
        self.running = True
        self.last_output_path = None
        self.recommendation_var.set("데이터를 확인해 분석 모드를 추천하는 중입니다.")

        # 분석 중에는 실행 버튼과 결과 폴더 버튼을 비활성화한다.
        self.run_button.configure(state="disabled")
        self.open_folder_button.configure(state="disabled")

        # 정확한 진행률 계산 대신 작업 중임을 보여주는 반복 진행바를 사용한다.
        self.progress.start(10)
        self.status_var.set("분석 중")

        self._log("=" * 68)
        self._log(f"{settings['analysis_mode']} 분석을 시작합니다.")
        self._log("=" * 68)

        # Word2Vec 학습은 시간이 오래 걸릴 수 있으므로 별도 스레드에서 실행한다.
        worker = threading.Thread(
            target=self._analysis_worker,
            args=(settings,),
            daemon=True,
        )
        worker.start()

    def _analysis_worker(self, settings: dict[str, object]) -> None:
        """
        실제 파일 읽기, 형태소 분석, 선택한 방식의 계산, Excel 저장을 수행한다.

        이 함수는 백그라운드 스레드에서 실행된다.
        messagebox나 tkinter 위젯을 직접 건드리지 않고 메시지 큐를 사용한다.
        """

        try:
            # -----------------------------------------------------------------
            # 1) 여러 TXT 파일 읽기
            # -----------------------------------------------------------------
            input_paths = list(settings["input_paths"])
            sentences, sentence_records = load_sentences(input_paths, self._log)

            # 단어별 품사 등장 횟수를 저장한다.
            token_pos_counts: dict[str, Counter] = defaultdict(Counter)

            alias_rules = list(settings["alias_rules"])
            query_words = list(settings["query_words"])
            requested_query_words = list(settings["requested_query_words"])

            if alias_rules:
                self._log("표현 통합 규칙 적용: " + str(settings["alias_rules_text"]))

            if requested_query_words != query_words:
                self._log(
                    "기준어 통합: "
                    + ", ".join(requested_query_words)
                    + " → "
                    + ", ".join(query_words)
                )

            # -----------------------------------------------------------------
            # 2) 기준어를 Kiwi 사용자 단어로 자동 등록
            # -----------------------------------------------------------------
            effective_user_words = list(settings["user_words"])
            registered_words = {
                parse_user_word(spec)[0]
                for spec in effective_user_words
            }

            for word in query_words:
                # 띄어쓰기가 없는 기준어는 하나의 고유명사로 자동 등록한다.
                if word not in registered_words and " " not in word:
                    effective_user_words.append(f"{word}:NNP:5.0")
                    registered_words.add(word)

            # -----------------------------------------------------------------
            # 3) 형태소 분석기 준비
            # -----------------------------------------------------------------
            self._log("형태소 분석기를 준비하는 중입니다.")

            tokenizer = build_tokenizer(
                user_words=effective_user_words,
                user_dict_path=settings["user_dict_path"],
                alias_rules=alias_rules,
                token_pos_counts=token_pos_counts,
                log=self._log,
            )

            if tokenizer is None:
                # 사용자 단어/사전 기능을 요청했는데 Kiwi가 없으면 정확한 실행이 불가능하다.
                if settings["user_words"] or settings["user_dict_path"]:
                    raise ImportError(
                        "사용자 단어 또는 사용자 사전을 사용하려면 "
                        "kiwipiepy가 필요합니다. "
                        "'pip install kiwipiepy'를 실행하세요."
                    )

                self._log(
                    "[경고] kiwipiepy가 없어 단순 문자열 기준으로 토큰화합니다. "
                    "정확한 한국어 분석을 위해 kiwipiepy 설치를 권장합니다."
                )

                tokenizer = build_simple_tokenizer(
                    alias_rules,
                    token_pos_counts,
                )
            else:
                self._log("Kiwi 형태소 분석기 준비 완료")

            # -----------------------------------------------------------------
            # 4) 전체 문장 토큰화
            # -----------------------------------------------------------------
            self._log("모든 파일의 문장을 형태소 단위로 분리하는 중입니다.")
            tokenized = tokenize_sentences(sentences, tokenizer)

            if len(tokenized) < 2:
                raise ValueError(
                    "형태소 분석 후 학습 가능한 문장이 2개 미만입니다. "
                    "입력 데이터와 품사 설정을 확인하세요."
                )

            self._log(f"학습 가능한 문장 수: {len(tokenized):,}개")

            # -----------------------------------------------------------------
            # 5) 전체 데이터 규모에 따른 분석 모드 추천
            # -----------------------------------------------------------------
            total_token_count = sum(len(sentence) for sentence in tokenized)
            recommended_mode = recommend_analysis_mode(total_token_count)
            recommendation_text = f"이 데이터에는 {recommended_mode}를 추천합니다."
            self.message_queue.put(("recommendation", recommendation_text))
            self._log(recommendation_text)

            # -----------------------------------------------------------------
            # 6) 선택한 방식으로 기준어 연관어 계산
            # -----------------------------------------------------------------
            analysis_mode = str(settings["analysis_mode"])
            model = None

            if analysis_mode == "PMI":
                self._log("PMI 연관어 계산을 시작합니다.")
                found_targets, result_rows = get_pmi_results(
                    tokenized_sentences=tokenized,
                    query_words=query_words,
                    top_n=int(settings["top_n"]),
                    window=int(settings["window"]),
                    min_count=int(settings["min_count"]),
                    excluded_words=set(settings["excluded_words"]),
                    result_mode=str(settings["result_mode"]),
                    token_pos_counts=token_pos_counts,
                    log=self._log,
                )
                method_name = "PMI"
                settings_text = (
                    f"방식=PMI, 통계=PMI·G², "
                    f"window={settings['window']}, "
                    f"min_count={settings['min_count']}, "
                    f"top_n={settings['top_n']}, "
                    "분석품사=명사·영문·숫자·동사·형용사·부사"
                )
            else:
                method_name = "Skip-gram" if settings["sg"] == 1 else "CBOW"

                self._log(
                    "Word2Vec 학습 시작: "
                    f"{method_name}, "
                    f"벡터 {settings['vector_size']}, "
                    f"문맥 {settings['window']}, "
                    f"반복 {settings['epochs']}"
                )

                model = train_word2vec(
                    tokenized_sentences=tokenized,
                    vector_size=int(settings["vector_size"]),
                    window=int(settings["window"]),
                    min_count=int(settings["min_count"]),
                    epochs=int(settings["epochs"]),
                    sg=int(settings["sg"]),
                )

                self._log("Word2Vec 학습 완료")
                found_targets, result_rows = get_similarity_results(
                    model=model,
                    query_words=query_words,
                    top_n=int(settings["top_n"]),
                    excluded_words=set(settings["excluded_words"]),
                    result_mode=str(settings["result_mode"]),
                    token_pos_counts=token_pos_counts,
                    log=self._log,
                )
                settings_text = (
                    f"방식={method_name}, "
                    f"vector_size={settings['vector_size']}, "
                    f"window={settings['window']}, "
                    f"min_count={settings['min_count']}, "
                    f"epochs={settings['epochs']}, "
                    f"top_n={settings['top_n']}, "
                    "학습품사=명사·영문·숫자·동사·형용사·부사"
                )

            if not result_rows:
                self._log("[경고] Excel에 기록할 연관어 결과가 없습니다.")

            # -----------------------------------------------------------------
            # 7) Excel 저장
            # -----------------------------------------------------------------
            self._log("Excel 결과를 저장하는 중입니다.")

            save_excel_report(
                output_path=str(settings["excel_path"]),
                sentence_records=sentence_records,
                input_file_names=list(settings["input_file_names"]),
                requested_targets=requested_query_words,
                normalized_targets=query_words,
                found_targets=found_targets,
                result_rows=result_rows,
                source=str(settings["source"]),
                settings_text=settings_text,
                alias_rules_text=str(settings["alias_rules_text"]),
                excluded_words_text=str(settings["excluded_words_text"]),
                result_mode=str(settings["result_mode"]),
                analysis_mode=analysis_mode,
            )

            self._log(f"Excel 저장 완료: {settings['excel_path']}")

            # -----------------------------------------------------------------
            # 8) 선택적 모델 저장
            # -----------------------------------------------------------------
            if analysis_mode == "Word2Vec" and settings["save_model"]:
                assert model is not None
                model_output = Path(str(settings["model_path"]))
                model_output.parent.mkdir(parents=True, exist_ok=True)
                model.save(str(model_output))
                self._log(f"모델 저장 완료: {model_output}")

            # -----------------------------------------------------------------
            # 9) 선택적 PCA 그래프 저장
            # -----------------------------------------------------------------
            if analysis_mode == "Word2Vec" and settings["save_plot"]:
                assert model is not None
                plot_embeddings(
                    model=model,
                    path=str(settings["plot_path"]),
                    target_words=found_targets,
                    result_rows=result_rows,
                    max_words=int(settings["plot_max_words"]),
                    log=self._log,
                )

            # 완료 신호를 GUI 스레드로 전달한다.
            self.message_queue.put(
                (
                    "done",
                    {
                        "path": str(settings["excel_path"]),
                        "analysis_mode": analysis_mode,
                    },
                )
            )

        except Exception as exc:
            # 오류가 발생하면 오류 메시지만 GUI 스레드로 전달한다.
            self.message_queue.put(("error", str(exc)))

    # -------------------------------------------------------------------------
    # 로그와 메시지 큐 처리
    # -------------------------------------------------------------------------
    def _log(self, message: str) -> None:
        """백그라운드 스레드에서도 안전하게 로그 메시지를 큐에 넣는다."""

        self.message_queue.put(("log", message))

    def _append_log(self, message: str) -> None:
        """로그 창에 한 줄을 추가한다. GUI 메인 스레드에서만 호출한다."""

        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def clear_log(self) -> None:
        """로그 창의 내용을 모두 지운다."""

        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _process_queue(self) -> None:
        """
        백그라운드 스레드가 보낸 메시지를 주기적으로 처리한다.

        메시지 종류
        ---------
        log   : 로그 창에 출력
        recommendation : 데이터 규모에 따른 추천 문구 표시
        done           : 분석 완료 처리
        error          : 오류 처리
        """

        try:
            while True:
                kind, payload = self.message_queue.get_nowait()

                if kind == "log":
                    self._append_log(str(payload))

                elif kind == "recommendation":
                    self.recommendation_var.set(str(payload))

                elif kind == "done":
                    result = dict(payload)
                    output_path = str(result["path"])
                    analysis_mode = str(result["analysis_mode"])
                    self.running = False
                    self.progress.stop()
                    self.status_var.set("완료")
                    self.run_button.configure(state="normal")

                    self.last_output_path = Path(output_path)
                    self.open_folder_button.configure(state="normal")

                    self._append_log("=" * 68)
                    self._append_log("모든 작업이 완료되었습니다.")

                    messagebox.showinfo(
                        "분석 완료",
                        f"{analysis_mode} 분석과 Excel 저장이 완료되었습니다.\n\n"
                        + output_path,
                    )

                elif kind == "error":
                    self.running = False
                    self.progress.stop()
                    self.status_var.set("오류 발생")
                    self.run_button.configure(state="normal")

                    self._append_log(f"[오류] {payload}")
                    messagebox.showerror("분석 실패", str(payload))

        except queue.Empty:
            # 처리할 메시지가 없으면 아무 작업도 하지 않는다.
            pass

        finally:
            # 다음 확인을 예약한다.
            self.root.after(100, self._process_queue)

    # -------------------------------------------------------------------------
    # 결과 폴더 열기
    # -------------------------------------------------------------------------
    def open_result_folder(self) -> None:
        """마지막으로 저장한 Excel 파일이 있는 폴더를 운영체제 탐색기로 연다."""

        if not self.last_output_path:
            return

        folder = self.last_output_path.resolve().parent

        try:
            if sys.platform.startswith("win"):
                # Windows 파일 탐색기
                os.startfile(folder)  # type: ignore[attr-defined]

            elif sys.platform == "darwin":
                # macOS Finder
                subprocess.run(["open", str(folder)], check=True)

            else:
                # Linux 계열 파일 관리자
                subprocess.run(["xdg-open", str(folder)], check=True)

        except Exception as exc:
            messagebox.showerror("폴더 열기 실패", str(exc))


# =============================================================================
# 9. 프로그램 시작점
# =============================================================================

def main() -> None:
    """tkinter 창을 만들고 GUI 이벤트 루프를 시작한다."""

    root = tk.Tk()

    # 변수에 저장하지 않아도 GUI 객체는 root가 참조하지만,
    # 명시적으로 app 변수에 보관하면 코드 구조가 더 분명하다.
    app = Word2VecGUI(root)

    # 일부 검사기에서 미사용 변수 경고가 나지 않도록 참조한다.
    _ = app

    # 창이 닫힐 때까지 사용자 입력과 이벤트를 처리한다.
    root.mainloop()


# 이 파일을 직접 실행했을 때만 main()을 호출한다.
# 다른 Python 파일에서 import할 때는 GUI가 자동 실행되지 않는다.
if __name__ == "__main__":
    main()
